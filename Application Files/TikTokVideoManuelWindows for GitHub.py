import json
import os
from datetime import datetime

import pyautogui
import undetected_chromedriver as uc
import time
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import locale

locale.setlocale(locale.LC_TIME, "tr_TR.UTF-8")

def Driver():
    options = uc.ChromeOptions()
    options.add_argument("--disable-popup-blocking")
    #options.add_argument("--incognito")
    #options.add_argument("--headless")  # Run the browser in headless mode
    options.add_argument("--start-maximized")  # Start the browser maximized
    driver = uc.Chrome(options=options)
    return driver

def Chrome_Tab(driver, link):
    driver.get(link)
    driver.maximize_window()

path = "<ExcelFilePath>"
# <ExcelFilePath>: Path to the Excel file
workbook = openpyxl.load_workbook(path)
DefaultPageSheet = workbook["<SheetName>"]
# <SheetName>: Name of the sheet in the Excel file

SocialMediaName = "TikTok"

StartColumn = 5
while True:
    Finder = DefaultPageSheet.cell(2, StartColumn).value
    if Finder == SocialMediaName:
        break
    StartColumn += 1

TotalAccountListGmail = []
DefaultRow = 4

while True:
    Start = DefaultPageSheet.cell(DefaultRow, StartColumn).value
    if Start is None:
        break
    if DefaultPageSheet.cell(DefaultRow, StartColumn+1).value != "-":
        TotalAccountListGmail.append([Start, DefaultRow - 3])
    DefaultRow += 1

# Filter out entries with '-'
TotalAccountListGmail = [item for item in TotalAccountListGmail if item[0] != '-']

TotalAccountList = []
DefaultRow = 4

for account_info in TotalAccountListGmail:
    index = account_info[1]
    Start = DefaultPageSheet.cell(index + 3, 2).value
    TotalAccountList.append(Start)

print(TotalAccountList)

# Get the starting index from the user
baslangic_indeksi = int(input("Enter the last completed index for TikTok accounts (0 to {}): ".format(len(TotalAccountList)-1)))

TotalAccountList = TotalAccountList[baslangic_indeksi:]
TotalAccountListGmail = TotalAccountListGmail[baslangic_indeksi:]

path = "<TikTokExcelFilePath>"
# <TikTokExcelFilePath>: Path to the TikTok Excel file
workbook = openpyxl.load_workbook(path)
AllPagesWorksheet = workbook.sheetnames

for IndexAccount, OneItem in enumerate(TotalAccountListGmail, start=0):
    driver = Driver()
    link = "https://www.tiktok.com"

    Chrome_Tab(driver, link)

    JsonFileName = DefaultPageSheet.cell(2, StartColumn).value + OneItem[0] + ".json"
    with open(f"<APIsPath>/{JsonFileName}", "r") as file:
        # <APIsPath>: Path to the APIs directory
        cookies = json.load(file)

    for cookie in cookies:
        driver.add_cookie(cookie)

    time.sleep(3)
    driver.refresh()

    StartRow = 5
    PostDict = {}

    OneAccountName = TotalAccountList[IndexAccount]
    AllPostList = []
    StartRow = 5
    while True:
        SpecialWorkSheet = workbook[OneAccountName]
        PostName = SpecialWorkSheet["B" + str(StartRow)].value
        StartRow += 1
        if PostName is None:
            break
        AllPostList.append(PostName)

    print(AllPostList)
    AllPostDefault = AllPostList
    baslangic_indeksi = int(input(f"Enter the last completed index for {OneAccountName} (0 to {(len(AllPostList) - 1)}): "))

    AllPostList = AllPostList[baslangic_indeksi:]

    for StartRow, PostOne in enumerate(AllPostDefault, start=5):
        SpecialWorkSheet = workbook[OneAccountName]
        PostName = SpecialWorkSheet["B" + str(StartRow)].value
        Aciklama = SpecialWorkSheet["C" + str(StartRow)].value

        Tarih = SpecialWorkSheet["D" + str(StartRow)].value
        Tarih = datetime.strptime(Tarih, '%d.%m.%Y')
        FormatliTarih = Tarih.strftime("%d %B %A %Y")
        Gun = Tarih.strftime("%d")
        Ay = Tarih.strftime("%B")
        Saat = str(SpecialWorkSheet["E" + str(StartRow)].value)
        SaatSplitted = Saat.split(":")
        Hour = str(SaatSplitted[0])
        Minutes = str(SaatSplitted[1])
        HemenPaylas = SpecialWorkSheet["F" + str(StartRow)].value

        post_details = {
            "Aciklama": Aciklama,
            "Tarih": FormatliTarih,
            "Gün": Gun,
            "Ay": Ay,
            "Saat": Saat,
            "Hour": Hour,
            "Minutes": Minutes,
            "HemenPaylas": HemenPaylas
        }

        if OneAccountName in PostDict:
            PostDict[OneAccountName][PostName] = post_details
        else:
            PostDict[OneAccountName] = {PostName: post_details}

    for StartRow, PostOne in enumerate(AllPostList, start=5):
        wait = WebDriverWait(driver, 60)
        sekmeler = driver.window_handles
        ilk_sekme = sekmeler[0]
        driver.switch_to.window(ilk_sekme)

        wait.until(EC.visibility_of_element_located((By.XPATH, "<UploadButton>"))).click()
        # <UploadButton>: XPath for upload button

        time.sleep(2)
        if StartRow == 5:
            for i in range(3):
                time.sleep(1)
                pyautogui.hotkey('ctrl', '-')
        time.sleep(1)

        time.sleep(5)
        iframe = driver.find_element(By.XPATH, "<IframeUpload>")
        # <IframeUpload>: XPath for iframe upload

        driver.switch_to.frame(iframe)
        FileAdress = "<TikTokVideoPath>/" + OneAccountName + "/" + PostOne + ".mp4"
        # <TikTokVideoPath>: Path to the TikTok videos
        driver.find_element(By.XPATH, "<FileInput>").send_keys(FileAdress)
        # <FileInput>: XPath for file input

        ProgressBarInvisibility = wait.until(EC.invisibility_of_element_located((By.XPATH, "<ProgressBar>")))
        # <ProgressBar>: XPath for progress bar
        time.sleep(20)
        try:
            while True:
                if str(driver.find_element(By.XPATH, "<ProgressText>").text) == "100%":
                    break
                time.sleep(5)
        except:
            wait = WebDriverWait(driver, 120)
            ProgressBarInvisibility = wait.until(EC.invisibility_of_element_located((By.XPATH, "<UploadingStageText>")))
            # <UploadingStageText>: XPath for uploading stage text
            time.sleep(1)

        time.sleep(1)

        FileAdressCover = "<TikTokCoverPath>/" + OneAccountName + "/" + PostOne + "-Kapak.jpg"
        # <TikTokCoverPath>: Path to the TikTok covers

        if os.path.exists(FileAdressCover):
            time.sleep(2)
            driver.find_element(By.XPATH, "<EditCoverButton>").click()
            # <EditCoverButton>: XPath for edit cover button
            time.sleep(2)
            driver.find_element(By.XPATH, "<UploadCoverButton>").click()
            # <UploadCoverButton>: XPath for upload cover button
            time.sleep(2)
            driver.find_element(By.XPATH, "<CoverInput>").send_keys(FileAdressCover)
            # <CoverInput>: XPath for cover input
            time.sleep(3)
            driver.find_element(By.XPATH, "<ConfirmButton>").click()
            # <ConfirmButton>: XPath for confirm button
            time.sleep(10)
        else:
            pass

        time.sleep(2)
        try:
            TextMessagexPath = "<TextMessageXpath>"
        except:
            TextMessagexPath = "<TextMessageXpathAlt>"
        # <TextMessageXpath>: XPath for text message
        # <TextMessageXpathAlt>: Alternative XPath for text message
        driver.find_element(By.XPATH, TextMessagexPath).send_keys(Keys.CONTROL+"a")
        driver.find_element(By.XPATH, TextMessagexPath).send_keys(str(PostDict[OneAccountName][PostOne]["Aciklama"]))

        time.sleep(2)
        try:
            PlanlaxPath = "<ScheduleButton>"
            driver.find_element(By.XPATH, PlanlaxPath).click()
        except:
            PlanlaxPath = "<ScheduleButtonAlt>"
            driver.find_element(By.XPATH, PlanlaxPath).click()
        # <ScheduleButton>: XPath for schedule button
        # <ScheduleButtonAlt>: Alternative XPath for schedule button

        time.sleep(1)
        driver.find_element(By.XPATH, "<DatePicker>").click()
        # <DatePicker>: XPath for date picker
        element = driver.find_element(By.XPATH, "<MonthTitle>").text
        # <MonthTitle>: XPath for month title

        if "May" in str(PostDict[OneAccountName][PostOne]["Ay"]):
            PostDict[OneAccountName][PostOne]["Ay"] = "May"

        time.sleep(1)
        if element == str(PostDict[OneAccountName][PostOne]["Ay"]):
            pass
        else:
            driver.find_element(By.XPATH, "<RightArrow>").click()
            # <RightArrow>: XPath for right arrow

        DateNo = str(PostDict[OneAccountName][PostOne]["Gün"])
        if "0" in DateNo[0]:
            DateNo = DateNo[1]

        for i in range(2):
            try:
                time.sleep(1)
                driver.find_element(By.XPATH, "<CalendarDay>[" + str(i+1) + "]").click()
                # <CalendarDay>: XPath for calendar day
            except:
                pass

        time.sleep(1)
        driver.find_element(By.XPATH, "<TimePicker>").click()
        # <TimePicker>: XPath for time picker

        HourElement = driver.find_element(By.XPATH, "<HourElementXpath>")
        # <HourElementXpath>: XPath for hour element
        MinutesElement = driver.find_element(By.XPATH, "<MinutesElementXpath>")
        # <MinutesElementXpath>: XPath for minutes element

        HourANDMinutes = [HourElement, MinutesElement]

        for HourORMinutes in HourANDMinutes:
            action = ActionChains(driver)
            time.sleep(2)
            action.move_to_element(HourORMinutes).perform()
            time.sleep(3)
            HourORMinutes.click()
            time.sleep(3)

        time.sleep(2)
        driver.find_element(By.XPATH, "<TimePicker>").click()
        # <TimePicker>: XPath for time picker
        time.sleep(1)
        try:
            driver.find_element(By.XPATH, "<MoreButton>").click()
            # <MoreButton>: XPath for more button
        except:
            pass

        time.sleep(2)
        driver.find_element(By.XPATH, "<AIContentCheckbox>").click()
        # <AIContentCheckbox>: XPath for AI content checkbox
        time.sleep(2)

        try:
            time.sleep(2)
            driver.find_element(By.XPATH, "<OpenButton>").click()
            # <OpenButton>: XPath for open button
        except:
            pass
        driver.find_element(By.XPATH, "<ScheduleButtonConfirm>").click()
        # <ScheduleButtonConfirm>: XPath for schedule button confirm
        time.sleep(2)
        driver.find_element(By.XPATH, "<ManagePostsButton>").click()
        # <ManagePostsButton>: XPath for manage posts button
        time.sleep(1)
        driver.switch_to.default_content()
        time.sleep(2)
        driver.get("https://www.tiktok.com/")
        time.sleep(3)
        print(f"{PostOne} completed")

    driver.quit()

print("Completed...")
